from openpyxl import Workbook
import datetime

wb = Workbook()

ws = wb.active

ws.title = "입력된 시간"

ws['A1'] = "초기위치값"

ws.append([10,20,30])

ws['B1'] = datetime.datetime.now()
ws['C1'] = "입력된 시간입니다."

wb.save("D:\\NowDate.xlsx")