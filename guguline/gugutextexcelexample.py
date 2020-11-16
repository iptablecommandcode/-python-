from openpyxl import load_workbook

# 변수선언 초기화
Date_Value = ""

wb = load_workbook("D:\\new.excel.xlsx")

ws2 = wb.create_sheet(title = "두번째 Sheet")

ws2.cell(1,5,"구구단을 엑셀로 만들었습니다.")


for row in range(2,10):
    f = open("D:\{}단.txt".format(row),'w')
    Date_Value = ("# %5d단 #\n"%row)
    f.write(Date_Value)

    Date_Value = ""

    # cell 사용
    for col in range(1,10):
        Date_Value = "{} X {} = {}".format(row,col,row*col)
        ws2.cell(row + 1, col, Date_Value)
        Date_Value = Date_Value + "\n"
        f.write(Date_Value)
        Date_Value = ""

f.close()
wb.save("D:\\test.xlsx")