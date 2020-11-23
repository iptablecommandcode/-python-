import requests

from bs4 import BeautifulSoup

from datetime import datetime as dt

from openpyxl import Workbook

wb = Workbook()

ws1 = wb.active

ws1.title = "1606053_Park ChiWon"

now = dt.now()

targetSite = 'https://www.melon.com/chart/index.htm'

print("{} 현재 Melon 실시간 차트 Top 100".format(now))

header = {'User-agent':'Mozilla/5.0(Windows NT 6.1: WOW64;Trident/7.0;rv:11.0) Like Gecko'}

request = requests.get(targetSite,headers = header)

html = request.text

soup = BeautifulSoup(html,'html.parser')

ranks = soup.findAll('div',{'class':'ellipsis rank01'})

artists = soup.findAll('span',{'class':'checkEllipsis'})

# print(ranks)

# print(artists)

for i in range(len(ranks)):
    artist = artists[i].text.strip().split('\n')[0]

    rank = ranks[i].text.strip().split('\n')[0]

    cell_number = '{0:3d}'.format(i+1)
    cell_title = '{}'.format(artist)
    cell_name = '{}'.format(rank)

    ws1.cell(i+1,1,cell_number)
    ws1.cell(i+1,2,cell_title)
    ws1.cell(i+1,3,cell_name)

wb.save("D:\\1606053_박치원.xlsx")